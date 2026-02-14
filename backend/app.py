from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import sqlite3
from datetime import datetime
import random
import os
from openpyxl import Workbook, load_workbook
from io import BytesIO

app = Flask(__name__)
CORS(app)

DATABASE = 'characters.db'

def get_db():
    """获取数据库连接"""
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    return db

def init_db():
    """初始化数据库"""
    db = get_db()
    cursor = db.cursor()
    
    # 创建汉字表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS characters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            character TEXT UNIQUE NOT NULL,
            recognition_count INTEGER DEFAULT 0,
            is_mastered BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 创建学习记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS learning_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            character_id INTEGER NOT NULL,
            recognized BOOLEAN NOT NULL,
            recorded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (character_id) REFERENCES characters(id)
        )
    ''')
    
    # 检查是否已有数据
    cursor.execute('SELECT COUNT(*) as count FROM characters')
    count = cursor.fetchone()['count']
    
    if count == 0:
        # 预制100个常用汉字
        common_characters = [
            '的', '一', '是', '在', '不', '了', '有', '和', '人', '这',
            '中', '大', '为', '上', '个', '国', '我', '以', '要', '他',
            '时', '来', '用', '们', '生', '到', '作', '地', '于', '出',
            '就', '分', '对', '成', '会', '可', '主', '发', '年', '动',
            '同', '工', '也', '能', '下', '过', '子', '说', '产', '种',
            '面', '而', '方', '后', '多', '定', '行', '学', '法', '所',
            '民', '得', '经', '十', '三', '之', '进', '着', '等', '部',
            '度', '家', '电', '力', '里', '如', '水', '化', '高', '自',
            '二', '理', '起', '小', '物', '现', '实', '加', '量', '都',
            '两', '体', '制', '机', '当', '使', '点', '从', '业', '本'
        ]
        
        for char in common_characters:
            cursor.execute(
                'INSERT INTO characters (character) VALUES (?)',
                (char,)
            )
    
    db.commit()
    db.close()

# API路由

@app.route('/api/characters', methods=['GET'])
def get_characters():
    """获取所有汉字"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        SELECT id, character, recognition_count, is_mastered, created_at, updated_at
        FROM characters
        ORDER BY created_at DESC
    ''')
    characters = [dict(row) for row in cursor.fetchall()]
    db.close()
    return jsonify(characters)

@app.route('/api/characters', methods=['POST'])
def add_character():
    """添加新汉字"""
    data = request.json
    character = data.get('character', '').strip()
    
    if not character or len(character) != 1:
        return jsonify({'error': '请输入单个汉字'}), 400
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute(
            'INSERT INTO characters (character) VALUES (?)',
            (character,)
        )
        db.commit()
        character_id = cursor.lastrowid
        
        cursor.execute('SELECT * FROM characters WHERE id = ?', (character_id,))
        new_character = dict(cursor.fetchone())
        db.close()
        
        return jsonify(new_character), 201
    except sqlite3.IntegrityError:
        db.close()
        return jsonify({'error': '该汉字已存在'}), 400

@app.route('/api/characters/batch', methods=['POST'])
def add_characters_batch():
    """批量添加汉字"""
    data = request.json
    characters_text = data.get('characters', '').strip()
    
    if not characters_text:
        return jsonify({'error': '请输入汉字'}), 400
    
    # 提取所有汉字
    chars = []
    for char in characters_text:
        if char.strip() and '\u4e00' <= char <= '\u9fff':  # 判断是否为汉字
            if char not in chars:  # 去重
                chars.append(char)
    
    if not chars:
        return jsonify({'error': '未找到有效的汉字'}), 400
    
    db = get_db()
    cursor = db.cursor()
    
    success_count = 0
    skip_count = 0
    
    for char in chars:
        try:
            cursor.execute(
                'INSERT INTO characters (character) VALUES (?)',
                (char,)
            )
            success_count += 1
        except sqlite3.IntegrityError:
            skip_count += 1
    
    db.commit()
    db.close()
    
    return jsonify({
        'success': success_count,
        'skipped': skip_count,
        'total': len(chars)
    }), 201

@app.route('/api/characters/<int:character_id>', methods=['GET'])
def get_character_detail(character_id):
    """获取汉字详情"""
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute('SELECT * FROM characters WHERE id = ?', (character_id,))
    character = cursor.fetchone()
    
    if not character:
        db.close()
        return jsonify({'error': '汉字不存在'}), 404
    
    db.close()
    return jsonify(dict(character))

@app.route('/api/characters/<int:character_id>', methods=['PUT'])
def update_character(character_id):
    """更新汉字详情"""
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    # 检查汉字是否存在
    cursor.execute('SELECT * FROM characters WHERE id = ?', (character_id,))
    character = cursor.fetchone()
    
    if not character:
        db.close()
        return jsonify({'error': '汉字不存在'}), 404
    
    # 更新字段
    pinyin = data.get('pinyin')
    definition = data.get('definition')
    words = data.get('words')
    sentences = data.get('sentences')
    
    cursor.execute('''
        UPDATE characters
        SET pinyin = ?, definition = ?, words = ?, sentences = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (pinyin, definition, words, sentences, character_id))
    
    db.commit()
    
    # 获取更新后的数据
    cursor.execute('SELECT * FROM characters WHERE id = ?', (character_id,))
    updated_character = dict(cursor.fetchone())
    db.close()
    
    return jsonify(updated_character)

@app.route('/api/characters/<int:character_id>', methods=['DELETE'])
def delete_character(character_id):
    """删除汉字"""
    db = get_db()
    cursor = db.cursor()
    
    # 先删除相关的学习记录
    cursor.execute('DELETE FROM learning_records WHERE character_id = ?', (character_id,))
    
    # 删除汉字
    cursor.execute('DELETE FROM characters WHERE id = ?', (character_id,))
    db.commit()
    db.close()
    
    return jsonify({'message': '删除成功'}), 200

@app.route('/api/characters/random', methods=['GET'])
def get_random_character():
    """获取随机的待学习汉字（默认未掌握的，复习模式获取已掌握的）"""
    mastered = request.args.get('mastered', 'false').lower() == 'true'

    db = get_db()
    cursor = db.cursor()

    if mastered:
        # 复习模式：获取已掌握的汉字
        cursor.execute('''
            SELECT id, character, recognition_count, is_mastered
            FROM characters
            WHERE is_mastered = 1
        ''')
        message = '没有已掌握的汉字可复习'
    else:
        # 学习模式：获取未掌握的汉字
        cursor.execute('''
            SELECT id, character, recognition_count, is_mastered
            FROM characters
            WHERE is_mastered = 0
        ''')
        message = '恭喜！所有汉字都已掌握'

    characters = [dict(row) for row in cursor.fetchall()]
    db.close()

    if not characters:
        return jsonify({'message': message}), 200

    # 随机选择一个
    random_char = random.choice(characters)
    return jsonify(random_char)

@app.route('/api/characters/<int:character_id>/mark', methods=['POST'])
def mark_character(character_id):
    """标记汉字认识或不认识"""
    data = request.json
    recognized = data.get('recognized', False)
    
    db = get_db()
    cursor = db.cursor()
    
    # 获取当前汉字信息
    cursor.execute('SELECT * FROM characters WHERE id = ?', (character_id,))
    character = cursor.fetchone()
    
    if not character:
        db.close()
        return jsonify({'error': '汉字不存在'}), 404
    
    # 记录学习记录
    cursor.execute(
        'INSERT INTO learning_records (character_id, recognized) VALUES (?, ?)',
        (character_id, recognized)
    )
    
    # 更新认识次数
    if recognized:
        new_count = character['recognition_count'] + 1
        is_mastered = 1 if new_count >= 3 else 0
        
        cursor.execute('''
            UPDATE characters
            SET recognition_count = ?, is_mastered = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (new_count, is_mastered, character_id))
    else:
        # 如果标记不认识，计数归零
        cursor.execute('''
            UPDATE characters
            SET recognition_count = 0, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (character_id,))
    
    db.commit()
    
    # 获取更新后的汉字信息
    cursor.execute('SELECT * FROM characters WHERE id = ?', (character_id,))
    updated_character = dict(cursor.fetchone())
    db.close()
    
    return jsonify(updated_character)

@app.route('/api/characters/<int:character_id>/reset', methods=['POST'])
def reset_character(character_id):
    """重置汉字学习进度"""
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute('''
        UPDATE characters
        SET recognition_count = 0, is_mastered = 0, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (character_id,))
    
    db.commit()
    
    cursor.execute('SELECT * FROM characters WHERE id = ?', (character_id,))
    updated_character = dict(cursor.fetchone())
    db.close()
    
    return jsonify(updated_character)

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取统计信息"""
    db = get_db()
    cursor = db.cursor()
    
    # 总字数
    cursor.execute('SELECT COUNT(*) as total FROM characters')
    total = cursor.fetchone()['total']
    
    # 已掌握字数
    cursor.execute('SELECT COUNT(*) as mastered FROM characters WHERE is_mastered = 1')
    mastered = cursor.fetchone()['mastered']
    
    # 学习中字数
    cursor.execute('SELECT COUNT(*) as learning FROM characters WHERE is_mastered = 0 AND recognition_count > 0')
    learning = cursor.fetchone()['learning']
    
    # 未开始字数
    cursor.execute('SELECT COUNT(*) as not_started FROM characters WHERE recognition_count = 0')
    not_started = cursor.fetchone()['not_started']
    
    # 今日学习记录
    cursor.execute('''
        SELECT COUNT(*) as today_count
        FROM learning_records
        WHERE DATE(recorded_at) = DATE('now', 'localtime')
    ''')
    today_count = cursor.fetchone()['today_count']
    
    # 今日认识的汉字
    cursor.execute('''
        SELECT DISTINCT c.character
        FROM learning_records lr
        JOIN characters c ON lr.character_id = c.id
        WHERE DATE(lr.recorded_at) = DATE('now', 'localtime') AND lr.recognized = 1
        ORDER BY lr.recorded_at DESC
    ''')
    today_recognized = [row['character'] for row in cursor.fetchall()]
    
    db.close()
    
    return jsonify({
        'total': total,
        'mastered': mastered,
        'learning': learning,
        'not_started': not_started,
        'today_count': today_count,
        'today_recognized': today_recognized,
        'progress': round((mastered / total * 100), 2) if total > 0 else 0
    })

@app.route('/api/records/recent', methods=['GET'])
def get_recent_records():
    """获取最近的学习记录"""
    limit = request.args.get('limit', 20, type=int)
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute('''
        SELECT lr.id, lr.recognized, lr.recorded_at, c.character
        FROM learning_records lr
        JOIN characters c ON lr.character_id = c.id
        ORDER BY lr.recorded_at DESC
        LIMIT ?
    ''', (limit,))
    
    records = [dict(row) for row in cursor.fetchall()]
    db.close()
    
    return jsonify(records)

@app.route('/api/mistakes', methods=['GET'])
def get_mistakes():
    """获取错题库 - 所有标记过不认识的汉字（排除已掌握和已认识的）"""
    db = get_db()
    cursor = db.cursor()
    
    # 查找所有有"不认识"记录的汉字
    # 排除条件：
    # 1. 已掌握的汉字 (is_mastered = 1)
    # 2. 有"认识"记录的汉字（已经标记过认识的）
    cursor.execute('''
        SELECT 
            c.id,
            c.character,
            c.recognition_count,
            c.is_mastered,
            c.pinyin,
            c.definition,
            c.words,
            c.sentences,
            COUNT(CASE WHEN lr.recognized = 0 THEN 1 END) as mistake_count,
            MAX(lr.recorded_at) as last_mistake_time
        FROM characters c
        INNER JOIN learning_records lr ON c.id = lr.character_id
        WHERE c.is_mastered = 0
        AND c.id NOT IN (
            SELECT DISTINCT character_id 
            FROM learning_records 
            WHERE recognized = 1
        )
        GROUP BY c.id
        HAVING mistake_count > 0
        ORDER BY last_mistake_time DESC
    ''')
    
    mistakes = [dict(row) for row in cursor.fetchall()]
    db.close()
    
    return jsonify(mistakes)

@app.route('/api/characters/export', methods=['GET'])
def export_characters_excel():
    """导出汉字库到Excel文件"""
    db = get_db()
    cursor = db.cursor()

    cursor.execute('''
        SELECT id, character, pinyin, definition, words, sentences,
               recognition_count, is_mastered, created_at, updated_at
        FROM characters
        ORDER BY character
    ''')

    characters = cursor.fetchall()
    db.close()

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "汉字库"

    # 设置表头
    headers = ['汉字', '拼音', '释义', '组词', '造句', '认识次数', '是否已掌握', '创建时间', '更新时间']
    ws.append(headers)

    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据
    for char in characters:
        ws.append([
            char['character'],
            char['pinyin'] or '',
            char['definition'] or '',
            char['words'] or '',
            char['sentences'] or '',
            char['recognition_count'],
            '是' if char['is_mastered'] else '否',
            char['created_at'],
            char['updated_at']
        ])

    # 调整列宽
    ws.column_dimensions['A'].width = 8  # 汉字
    ws.column_dimensions['B'].width = 15  # 拼音
    ws.column_dimensions['C'].width = 30  # 释义
    ws.column_dimensions['D'].width = 40  # 组词
    ws.column_dimensions['E'].width = 50  # 造句
    ws.column_dimensions['F'].width = 10  # 认识次数
    ws.column_dimensions['G'].width = 12  # 是否已掌握
    ws.column_dimensions['H'].width = 20  # 创建时间
    ws.column_dimensions['I'].width = 20  # 更新时间

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名（包含时间戳）
    filename = f"汉字库导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/characters/import', methods=['POST'])
def import_characters_excel():
    """从Excel文件导入汉字库"""
    if 'file' not in request.files:
        return jsonify({'error': '请上传Excel文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '请上传Excel文件'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传Excel文件格式'}), 400

    try:
        # 加载Excel文件
        wb = load_workbook(filename=file)
        ws = wb.active

        db = get_db()
        cursor = db.cursor()

        success_count = 0
        update_count = 0
        skip_count = 0
        error_list = []

        # 跳过表头，从第二行开始
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            try:
                # 解析行数据
                # 格式: [汉字, 拼音, 释义, 组词, 造句, 认识次数, 是否已掌握, 创建时间, 更新时间]
                character = str(row[0]).strip()

                if not character:
                    continue

                pinyin = str(row[1]).strip() if len(row) > 1 and row[1] else None
                definition = str(row[2]).strip() if len(row) > 2 and row[2] else None
                words = str(row[3]).strip() if len(row) > 3 and row[3] else None
                sentences = str(row[4]).strip() if len(row) > 4 and row[4] else None
                recognition_count = int(row[5]) if len(row) > 5 and row[5] is not None else 0
                is_mastered_str = str(row[6]).strip() if len(row) > 6 and row[6] else '否'
                is_mastered = 1 if is_mastered_str in ['是', 'Yes', 'TRUE', '1', 'true'] else 0

                # 检查汉字是否存在
                cursor.execute('SELECT id FROM characters WHERE character = ?', (character,))
                existing = cursor.fetchone()

                if existing:
                    # 更新现有汉字
                    cursor.execute('''
                        UPDATE characters
                        SET pinyin = ?, definition = ?, words = ?, sentences = ?,
                            recognition_count = ?, is_mastered = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                    ''', (pinyin, definition, words, sentences, recognition_count, is_mastered, existing['id']))
                    update_count += 1
                else:
                    # 插入新汉字
                    cursor.execute('''
                        INSERT INTO characters
                        (character, pinyin, definition, words, sentences, recognition_count, is_mastered)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (character, pinyin, definition, words, sentences, recognition_count, is_mastered))
                    success_count += 1

            except Exception as e:
                skip_count += 1
                error_list.append(f'第{idx}行: {str(e)}')

        db.commit()
        db.close()

        return jsonify({
            'success': success_count,
            'updated': update_count,
            'skipped': skip_count,
            'errors': error_list[:10]  # 只返回前10个错误
        }), 201

    except Exception as e:
        return jsonify({'error': f'文件处理失败: {str(e)}'}), 400

if __name__ == '__main__':
    # 初始化数据库
    if not os.path.exists(DATABASE):
        init_db()
        print('数据库初始化完成！')

    print('后端服务启动在 http://localhost:5000')
    print('局域网访问地址：http://<本机IP>:5000')
    app.run(debug=True, host='0.0.0.0', port=5000)
